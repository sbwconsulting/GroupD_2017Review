#create the data files for the report exhibits

import logging

from sqlalchemy.sql import select
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
import sqlalchemy as db
import pandas as pd
import locale
from locale import atof
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from cpuc.db import engine
from cpuc.mylogging import create_logfile
import cpuc.params as params
from cpuc.models import Base
from cpuc.models import Measure
from cpuc.models import Sample
from cpuc.models import Study
from cpuc.workbookfunctions import convertwstodf
from cpuc.workbookfunctions import openworkbook
from cpuc.workbookfunctions import getSampleControlFile
from cpuc.stepchanges import createSteps

Session = sessionmaker(bind=engine)

class PivotOutput:
    def __init__(self):
        self.sourcefile = ''
        self.outputfile = ''
        self.datasheet = ''
        self.pivotsheet = ''
        self.data = None

def generatefiles(session):
    #get field lists
    xreffile = params.D0_FIELD_CROSS_REF_FILE
    wb = openworkbook(xreffile)
    ws_map = wb['RptOutput']
    #convert to datframe
    df_map = convertwstodf(ws_map, 1)
 
    ordercol = 'FieldOrder'    
    claimcol = 'claimfields'
    clmfullcol = 'claimfull'
    dbrvwcol = 'dbreviewfields'
    #dbextcol = 'dbrvwextra'    
    evalcol = 'evalfields'
    atrcol = 'atrfields'
    stepcol = 'stepeqn'
    rawcol = 'rawextra'

    #get sampleframe
    df_frame = pd.read_csv(params.D0_DATA_PATH + '\\cust_trkg_data_2017.csv')
    #df_frame = df_frame.rename(columns={'Unamed: 0': 'excelcounter'})
    #df_frame = df_frame.set_index('ClaimID')
    cols = ['ClaimID', 'SBW.ProjID']
    df_frame_short = df_frame[cols]
    print (f'short type is {type(df_frame_short)}')
    #get claim data
    df_claim_full = pd.read_csv(params.D0_DATA_FILE, low_memory=False)
    df_claim_full = df_claim_full.merge(df_frame_short, on='ClaimID')
    # Exclude replaced measurements
    df_claim = df_claim_full[df_claim_full.Replaced != 'Yes']
    # Pick only sampled rows
    df_claim = df_claim[df_claim.sampled == 'Y']
    #limit to claim fields
    df_map_claimfields = df_map[df_map[claimcol].notnull()]
    df_map_claimfields = df_map_claimfields[[ordercol, claimcol]]
    df_claim = df_claim[df_map_claimfields[claimcol].tolist()]    
    claimdict = df_map_claimfields.set_index(claimcol)[ordercol].to_dict()
    df_claim = df_claim.rename(columns=claimdict)
    #print(f'claimshape is {df_claim.shape} with cols {df_claim.columns}')
        
    #get ATR claim data
    df_atr = pd.read_csv(params.D0_LOCAL_ATR_OUTPUT_FILE, low_memory=False)
    df_map_atrfields = df_map[df_map[atrcol].notnull()]
    df_map_atrfields = df_map_atrfields[[ordercol, atrcol]]
    df_atr = df_atr[df_map_atrfields[atrcol].tolist()]    
    atrdict = df_map_atrfields.set_index(atrcol)[ordercol].to_dict()
    df_atr = df_atr.rename(columns=atrdict)

    #get eval data
    measures_without_sample_id = pd.read_sql(session.query(Measure).statement, session.bind)
    #drop project fields
    smplfields = ['RvwInstallDate', 'RvwAppVsInstallDate', 'RvwPaidIncentive', 'RvwPermit']
    measures_without_sample_id.drop(smplfields, axis=1, inplace=True)    
    samples = pd.read_sql(session.query(Sample).statement, session.bind)
    measures1 = measures_without_sample_id.merge(samples, on='SBW_ProjID')
    df_map_evalfields = df_map[df_map[evalcol].notnull()]
    df_map_evalfields = df_map_evalfields[[ordercol, evalcol]]
    df_dups = None
    df_dups = df_map_evalfields[df_map_evalfields.duplicated(evalcol)]
    print('dup is {}'.format(df_dups))
    df_map_evalfields.drop(df_dups[ordercol].tolist(), axis=0, inplace=True)
    evaldict = df_map_evalfields.set_index(evalcol)[ordercol].to_dict()   
    measures = measures1[df_map_evalfields[evalcol].tolist()]
    measures = measures.rename(columns=evaldict)
    measures = measures.assign(SampledProject = 1)

    if len(df_dups.index) >0:
        #only works if there are not multiple of the same dup
        #add the dup fields
        evaldict = df_dups.set_index(evalcol)[ordercol].to_dict()
        tmp = measures1.rename(columns=evaldict)
        measures = measures.join(tmp[df_dups[ordercol].tolist()])
        del tmp

    df_eval = measures.merge(df_claim, on='ClaimID')
    df_eval = df_eval.merge(df_atr, on='ClaimID', suffixes=('eval', ''))
    print(f'df_eval shape with claim and atr is {df_eval.shape}')

    #Get data from alldata
    df_raw_extra = pd.read_csv(params.D0_ALL_DATA_FILE, low_memory=False)
    df_raw_extra = remapdata(df_map, rawcol, ordercol, df_raw_extra)
    df_eval = df_eval.merge(df_raw_extra, on='ClaimID', suffixes=('_old', '')) #shouldn't be any overlap, but just in case want to keep a clean set.

    #get db review data
    df_dbreview = pd.read_csv(params.D0_DATABASE_REVIEW_FILE)
    #df_dbreview_ext = pd.read_csv(params.D0_DATABASE_REVIEW_EXTENDED_FILE)
    
    #drop passthru records
    df_dbreview = df_dbreview.query('EvalStdReportGroup == "2017_Savings_Review"')
    df_map_dbrvwfields = df_map[df_map[dbrvwcol].notnull()]
    df_map_dbrvwfields = df_map_dbrvwfields[[ordercol, dbrvwcol]]
    df_dbreview = df_dbreview[df_map_dbrvwfields[dbrvwcol].tolist()]
    dbrwvdict = df_map_dbrvwfields.set_index(dbrvwcol)[ordercol].to_dict()
    df_dbreview = df_dbreview.rename(columns=dbrwvdict)
    '''
    #add extra fields
    df_map_dbextfields = df_map[df_map[dbextcol].notnull()]
    df_map_dbextfields = df_map_dbextfields[[ordercol, dbextcol]]
    df_dbreview_ext = df_dbreview_ext[df_map_dbextfields[dbextcol].tolist()]
    dbrwvdict = df_map_dbextfields.set_index(dbextcol)[ordercol].to_dict()
    df_dbreview_ext = df_dbreview_ext.rename(columns=dbrwvdict)
    df_dbreview = df_dbreview.merge(df_dbreview_ext, on='ClaimID')
    '''

    #add claimfull
    df_map_clmfullfields = df_map[df_map[clmfullcol].notnull()]
    df_map_clmfullfields = df_map_clmfullfields[[ordercol, clmfullcol]]
    df_claim_ext = df_claim_full[df_map_clmfullfields[clmfullcol].tolist()]
    cfulldict = df_map_clmfullfields.set_index(clmfullcol)[ordercol].to_dict()
    df_claim_ext = df_claim_ext.rename(columns=cfulldict)
    df_dbreview = df_dbreview.merge(df_claim_ext, on='ClaimID')

    #add source flag
    df_dbreview = df_dbreview.assign(ClaimsDatabase = 1)

    #the merge below appends dbreview rows to sample rows. Total rows = dbrevoew + sample
    #df_evalallrows = df_eval.append(df_dbreview)
    #the merge below creates extrs fields whey they overlap based on matching claimIDs. Total rows = dbreview rows.
    df_eval = df_eval.merge(df_dbreview, on='ClaimID', how='outer', suffixes=('', '_cdr'))
    
    #now add in the atr fields for the fullset
    df_eval = df_eval.merge(df_atr, on='ClaimID', how='left', suffixes=('', '_atr'))

    #Add in the steps for table 5
    df_steps = createSteps(df_eval)
    df_steps = df_steps.reset_index()
    df_steps = remapdata(df_map, stepcol, ordercol, df_steps)
    df_eval = df_eval.merge(df_steps, on='ClaimID' ,suffixes=('_old', '')) #shouldn't be any overlap, but just in case want to keep a clean set.
    
    # just a diagnostic section, I think
    ws_map = wb['Eligibility']
    #convert to datframe
    df_map_elig = convertwstodf(ws_map, 1)
    anycol = 'noteligfieldsany'
    allcol = 'noteligfieldsall'
    df_map_fields = df_map_elig[df_map_elig[anycol].notnull()]
    anyfields = df_map_fields[anycol].tolist()
    df_map_fields = df_map_elig[df_map_elig[allcol].notnull()]
    allfields = df_map_fields[allcol].tolist()
    tmp = ((df_eval[anyfields] == 'No').any(1) | (df_eval[allfields] == 'No').all(1))
    print(f'ineligible list length {(tmp.sum())}')

    t6cnts = (df_eval[anyfields] == 'No').sum()
    csvfile = params.SAMPLED_SITE_REVIEW_PATH + '\\settozerocnts.csv'
    t6cnts.to_csv(csvfile)
    #print(f't6? {t6cnts}')
    #end of mystery section. Is it just for diagnostic purposes?

    print(f'df_eval shape after dbreview merge is {df_eval.shape}')
    #print(f'df_evalallrows shape after dbreview append is {df_evalallrows.shape}')
    #flag for changed NTG and EUL IDs
    #df_eval = df_eval.assign(NTGIDChanged = df_eval['cdrNTG_ID'] != df_eval['EvalNTG_ID'])
    df_eval = df_eval.assign(EvalEUL_ID_trim = df_eval['EvalEUL_ID'].str.replace('_Any', '').str.replace('_Gro', ''))
    #df_eval = df_eval.assign(EULIDChanged = df_eval['cdrEUL_ID'] != df_eval['EvalEUL_ID_trim'])
    df_eval = df_eval.assign(MeasDescChanged = df_eval['EvalMeasDescription'] != df_eval['MeasDescription'])
    #new fields to indicate savings changed
    df_eval = df_eval.assign(kWhChanged = df_eval['EvalExPostLifecycleNetkWh_atr'] != df_eval['ExAnteLifecycleNetkWh'])
    df_eval = df_eval.assign(kWhPctChange = ((df_eval['EvalExPostLifecycleNetkWh_atr'] - df_eval['ExAnteLifecycleNetkWh'])/df_eval['ExAnteLifecycleNetkWh']))
    df_eval = df_eval.assign(thmChanged = df_eval['EvalExPostLifecycleNetTherm_atr'] != df_eval['ExAnteLifecycleNetTherm'])
    df_eval = df_eval.assign(thmPctChange = ((df_eval['EvalExPostLifecycleNetTherm_atr'] - df_eval['ExAnteLifecycleNetTherm'])/df_eval['ExAnteLifecycleNetTherm']))

    if True:
        #set final column order
        df_map_fields = df_map[df_map[ordercol].notnull()]
        #fieldorder = df_map_fields[ordercol].tolist()
        #df_final = df_eval[fieldorder]
        df_final = df_eval

        try:
            csvfile = params.D0_REPORT_DATA_FILE
            df_final.to_csv(csvfile, index=False)
            logging.info(f'wrote csv for {csvfile}')
        except:
            logging.info("file in use writing to backup")
            df_final.to_csv(params.SAMPLED_SITE_REVIEW_PATH + '\\evaldata2.csv', index=False)    

    #generate xl file
    exportdata = PivotOutput()
    exportdata.sourcefile = params.SAMPLED_SITE_REVIEW_PATH + '\\pivottemplate.xlsx'
    exportdata.outputfile = params.SAMPLED_SITE_REVIEW_PATH + '\\evaldata.xlsx'
    exportdata.datasheet = 'data'
    exportdata.pivotsheet = 'pvt_data'
    exportdata.data = df_final
    output_to_excel_pivot(exportdata)
        
    print('data exported')

    logging.info('Done generating report data files')

def output_to_excel_pivot(dataobj):
    wb = openworkbook(dataobj.sourcefile)
    datasheet = dataobj.datasheet
    ws_pvt = wb[dataobj.pivotsheet]
    ws_data = wb[datasheet]

    #clear data sheet and put in new data
    wb.remove(ws_data)
    wb.create_sheet(datasheet)
    ws_data = wb[datasheet]
    #header = ['id']
    #header.extend(dataobj.data.columns)
    #print(f'header type {type(header)} : {header}')
    #no idea why line above doesn't work but line below does. Output appears identical
    #header = ['id'] + [w.replace('i', 'i') for w in list(dataobj.data.columns)]
    header = [w.replace('i', 'i') for w in list(dataobj.data.columns)]
    #print(f'header after replace type {type(header)} : {header}')
    ws_data.append(header)
    for r in dataframe_to_rows(dataobj.data, index=False, header=False):
       ws_data.append(r)

    #update pivot data area
    pivot = ws_pvt._pivots[0]
    pivot.cache.cacheSource.worksheetSource.ref = f'A1:{get_column_letter(len(dataobj.data.columns))}{len(dataobj.data.index)+1}'
    pivot.cache.refreshOnload = True
    try:
        wb.save(filename=dataobj.outputfile)
    except:
        print('excel file in use. No output file made')
        logging.warning(f'file in use {dataobj.outputfile}')

def remapdata(df_map, srccol, destcol, df_src):
    #return the remaped src df
    df_map_fields = df_map[df_map[srccol].notnull()]
    df_map_fields = df_map_fields[[destcol, srccol]]
    df_new = df_src[df_map_fields[srccol].tolist()]
    fielddict = df_map_fields.set_index(srccol)[destcol].to_dict()
    df_new = df_new.rename(columns=fielddict)
    return df_new

def rungeneratefiles():
    session = Session()
    generatefiles(session)    
    logging.info('done with rungeneratefiles')

if __name__ == '__main__':
    create_logfile('../logs/createreportdata.log')
    logging.info('Beginning session')
    session = Session()
    generatefiles(session)
